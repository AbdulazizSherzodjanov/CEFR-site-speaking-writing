from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User
from django.utils.html import format_html
from django.http import HttpResponse


# ── Inject custom CSS into admin ───────────────────────────────────────────
class SpeakProAdminSite(admin.AdminSite):
    site_header  = "SpeakPro — CEFR Platform"
    site_title   = "SpeakPro Admin"
    index_title  = "Platform Management"

    class Media:
        css = {'all': ('admin/custom_admin.css',)}

# Override default admin site header/title (simpler approach)
admin.site.site_header  = "SpeakPro — CEFR Platform"
admin.site.site_title   = "SpeakPro Admin"
admin.site.index_title  = "📊 Platform Management"

from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Teacher, StudentProfile,
    Part11Question, Part12Group, Part12Question,
    Part2Question, Part3Question,
    MockTest, Candidate, TestSession, QuestionResponse, SiteAnalytics,
    Announcement, AnnouncementDismissal,
)


# ── Student Profile Admin ─────────────────────────────────────────────────

def approve_students(modeladmin, request, queryset):
    from datetime import date
    updated = queryset.filter(is_approved=False).update(
        is_approved=True,
        is_blocked=False,
        access_start=date.today(),
    )
    modeladmin.message_user(request, f"{updated} student(s) approved.")
approve_students.short_description = "Approve selected students (sets access start to today)"

def block_students(modeladmin, request, queryset):
    queryset.update(is_blocked=True)
    modeladmin.message_user(request, f"{queryset.count()} student(s) blocked.")
block_students.short_description = "Block selected students"

def unblock_students(modeladmin, request, queryset):
    queryset.update(is_blocked=False)
    modeladmin.message_user(request, f"{queryset.count()} student(s) unblocked.")
unblock_students.short_description = "Unblock selected students"


@admin.register(StudentProfile)
class StudentProfileAdmin(admin.ModelAdmin):
    list_display = (
        'get_full_name', 'get_username', 'teacher',
        'is_approved', 'is_blocked', 'access_start', 'access_days',
        'get_expiry', 'get_access_status', 'created_at',
    )
    list_editable = ('is_approved', 'is_blocked', 'access_days')
    list_filter = ('is_approved', 'is_blocked', 'teacher')
    search_fields = ('user__first_name', 'user__last_name', 'user__username')
    actions = [approve_students, block_students, unblock_students]
    readonly_fields = ('created_at', 'streak', 'total_tests', 'get_access_status')
    fieldsets = (
        ('Student Info', {
            'fields': ('user', 'teacher', 'streak', 'total_tests', 'created_at')
        }),
        ('Access Control', {
            'fields': ('is_approved', 'is_blocked', 'access_start', 'access_days', 'get_expiry', 'get_access_status'),
            'description': 'Approve student, set duration and daily test limits.'
        }),
    )

    def get_full_name(self, obj):
        return obj.user.get_full_name() or '—'
    get_full_name.short_description = 'Full Name'

    def get_username(self, obj):
        return obj.user.username
    get_username.short_description = 'Username'

    def get_expiry(self, obj):
        d = obj.expiry_date()
        if not d:
            return '∞ Unlimited'
        from django.utils import timezone
        if timezone.now().date() > d:
            return format_html('<span style="color:red">{} (expired)</span>', d.strftime('%b %d, %Y'))
        return d.strftime('%b %d, %Y')
    get_expiry.short_description = 'Expires'

    def get_access_status(self, obj):
        _, reason = obj.has_access()
        colors = {
            'ok': ('green', 'Active'),
            'pending': ('orange', 'Pending Approval'),
            'blocked': ('red', 'Blocked'),
            'expired': ('gray', 'Expired'),
        }
        color, label = colors.get(reason, ('gray', reason))
        return format_html('<span style="color:{};font-weight:bold">{}</span>', color, label)
    get_access_status.short_description = 'Status'

    def save_model(self, request, obj, form, change):
        # Auto-set access_start when approving for the first time
        if obj.is_approved and not obj.access_start:
            from datetime import date
            obj.access_start = date.today()
        super().save_model(request, obj, form, change)


# ── User + Profile ─────────────────────────────────────────────────────────

class StudentProfileInline(admin.StackedInline):
    model = StudentProfile
    can_delete = False
    fields = ('teacher', 'is_approved', 'is_blocked', 'access_start', 'access_days')
    readonly_fields = ()


class CustomUserAdmin(UserAdmin):
    inlines = (StudentProfileInline,)
    list_display = ('username', 'get_full_name', 'email', 'get_teacher', 'get_access_status', 'is_active')

    def get_teacher(self, obj):
        try:
            return obj.student_profile.teacher.name if obj.student_profile.teacher else '—'
        except Exception:
            return '—'
    get_teacher.short_description = 'Teacher'

    def get_access_status(self, obj):
        try:
            _, reason = obj.student_profile.has_access()
            colors = {
                'ok': ('green', 'Active'),
                'pending': ('orange', 'Pending'),
                'blocked': ('red', 'Blocked'),
                'expired': ('gray', 'Expired'),
            }
            color, label = colors.get(reason, ('gray', reason))
            return format_html('<span style="color:{};font-weight:bold">{}</span>', color, label)
        except Exception:
            return '—'
    get_access_status.short_description = 'Access'


admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)


# ── Teacher ────────────────────────────────────────────────────────────────

@admin.register(Teacher)
class TeacherAdmin(admin.ModelAdmin):
    list_display = ('name', 'telegram_id', 'email', 'is_active')
    list_editable = ('is_active',)
    search_fields = ('name', 'email')


# ── Part 1.1 — Plain text questions ───────────────────────────────────────

@admin.register(Part11Question)
class Part11QuestionAdmin(admin.ModelAdmin):
    list_display = ('question_number', 'text_preview', 'prep_time_seconds',
                    'answer_time_seconds', 'is_active', 'order')
    list_editable = ('is_active', 'order', 'prep_time_seconds', 'answer_time_seconds')
    list_filter = ('is_active',)
    search_fields = ('text',)

    def text_preview(self, obj):
        return obj.text[:90] + '…' if len(obj.text) > 90 else obj.text
    text_preview.short_description = 'Question Text'


# ── Part 1.2 — Group + image questions ────────────────────────────────────

class Part12QuestionInline(admin.TabularInline):
    model = Part12Question
    fields = ('question_number', 'text', 'prep_time_seconds', 'answer_time_seconds', 'is_active', 'order')
    extra = 3
    ordering = ('order', 'question_number')


@admin.register(Part12Group)
class Part12GroupAdmin(admin.ModelAdmin):
    list_display = ('title', 'image_preview', 'question_count', 'is_active', 'order')
    list_editable = ('is_active', 'order')
    inlines = [Part12QuestionInline]

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="height:40px;border-radius:6px">', obj.image.url)
        return '—'
    image_preview.short_description = 'Image'

    def question_count(self, obj):
        return obj.questions.count()
    question_count.short_description = 'Questions'


@admin.register(Part12Question)
class Part12QuestionAdmin(admin.ModelAdmin):
    list_display = ('question_number', 'group', 'text_preview', 'is_active')
    list_filter = ('group', 'is_active')

    def text_preview(self, obj):
        return obj.text[:80] + '…' if len(obj.text) > 80 else obj.text
    text_preview.short_description = 'Question'


# ── Part 2 ─────────────────────────────────────────────────────────────────

@admin.register(Part2Question)
class Part2QuestionAdmin(admin.ModelAdmin):
    list_display = ('question_number', 'text_preview', 'image_preview',
                    'prep_time_seconds', 'answer_time_seconds', 'is_active', 'order')
    list_editable = ('is_active', 'order', 'prep_time_seconds', 'answer_time_seconds')

    def text_preview(self, obj):
        return obj.text[:80] + '…' if len(obj.text) > 80 else obj.text
    text_preview.short_description = 'Prompt'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="height:40px;border-radius:6px">', obj.image.url)
        return '—'
    image_preview.short_description = 'Image'


# ── Part 3 ─────────────────────────────────────────────────────────────────

@admin.register(Part3Question)
class Part3QuestionAdmin(admin.ModelAdmin):
    list_display = ('question_number', 'text_preview', 'image_preview',
                    'prep_time_seconds', 'answer_time_seconds', 'is_active', 'order')
    list_editable = ('is_active', 'order', 'prep_time_seconds', 'answer_time_seconds')

    def text_preview(self, obj):
        return obj.text[:80] + '…' if len(obj.text) > 80 else obj.text
    text_preview.short_description = 'Task'

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="height:40px;border-radius:6px">', obj.image.url)
        return '—'
    image_preview.short_description = 'Image'


# ── Mock Test ──────────────────────────────────────────────────────────────

class CandidateInline(admin.TabularInline):
    model = Candidate
    fields = ('candidate_id', 'full_name', 'has_taken_test')
    extra = 3
    readonly_fields = ('has_taken_test',)


@admin.register(MockTest)
class MockTestAdmin(admin.ModelAdmin):
    list_display = ('title', 'code', 'part', 'teacher', 'is_active',
                    'valid_from', 'valid_until', 'candidate_count', 'created_at')
    list_filter = ('is_active', 'part', 'teacher')
    search_fields = ('title', 'code')
    inlines = [CandidateInline]
    readonly_fields = ('created_at',)

    def candidate_count(self, obj):
        return obj.candidates.count()
    candidate_count.short_description = 'Candidates'


@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'candidate_id', 'mock_test', 'has_taken_test', 'created_at')
    list_filter = ('mock_test', 'has_taken_test')
    search_fields = ('full_name', 'candidate_id')


# ── Test Session — with Excel export ──────────────────────────────────────

class QuestionResponseInline(admin.StackedInline):
    model = QuestionResponse
    readonly_fields = ('question_number', 'question_text', 'audio_player',
                       'transcription', 'score', 'score_breakdown', 'feedback',
                       'answered_at', 'duration_seconds')
    extra = 0
    can_delete = False

    def audio_player(self, obj):
        if obj.audio_file:
            try:
                return format_html(
                    '<audio controls style="width:320px">'
                    '<source src="{}" type="audio/webm">'
                    '</audio>', obj.audio_file.url)
            except Exception:
                return '(file missing)'
        return '(deleted)'
    audio_player.short_description = 'Audio'


def export_sessions_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    altfill = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")

    headers = ['#', 'Full Name', 'Type', 'Part', 'Teacher', 'Status',
               'Total Score', 'Started', 'Completed', 'Duration (min)',
               'Q1 Score', 'Q1 Transcript',
               'Q2 Score', 'Q2 Transcript',
               'Q3 Score', 'Q3 Transcript',
               'Telegram Sent', 'Audio Status']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.border = border
        cell.alignment = center

    for row_num, session in enumerate(queryset, 2):
        responses = list(session.responses.order_by('question_number'))
        duration = None
        if session.completed_at and session.started_at:
            duration = round((session.completed_at - session.started_at).total_seconds() / 60, 1)

        row = [
            row_num - 1,
            session.full_name,
            session.get_session_type_display(),
            f"Part {session.part}",
            session.teacher.name if session.teacher else '—',
            session.get_status_display(),
            round(session.total_score, 2) if session.total_score else '—',
            session.started_at.strftime('%Y-%m-%d %H:%M') if session.started_at else '—',
            session.completed_at.strftime('%Y-%m-%d %H:%M') if session.completed_at else '—',
            duration or '—',
        ]
        for i in range(3):
            if i < len(responses):
                r = responses[i]
                row.extend([r.score or '—', (r.transcription or '')[:300]])
            else:
                row.extend(['—', '—'])
        row.append('Yes' if session.telegram_sent else 'No')
        row.append('Deleted' if session.audio_deleted else 'Available')

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = altfill

    widths = [4, 22, 12, 10, 18, 12, 12, 18, 18, 12, 8, 45, 8, 45, 8, 45, 12, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'CEFR Speaking — Results Summary'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'] = f'Exported: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    ws2['A4'] = f'Total Sessions: {queryset.count()}'
    completed = queryset.filter(status='completed').count()
    ws2['A5'] = f'Completed: {completed}'
    ws2['A6'] = f'Completion Rate: {round(completed / queryset.count() * 100, 1) if queryset.count() else 0}%'
    scores = [s.total_score for s in queryset if s.total_score]
    ws2['A7'] = f'Average Score: {round(sum(scores) / len(scores), 2) if scores else "N/A"}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=cefr_results_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(response)
    return response


export_sessions_xlsx.short_description = "📊 Export selected to Excel (.xlsx)"


@admin.register(TestSession)
class TestSessionAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'session_type', 'part', 'teacher', 'status',
                    'total_score', 'started_at', 'telegram_sent', 'audio_status')
    list_filter = ('session_type', 'status', 'part', 'teacher', 'telegram_sent',
                   'audio_deleted', 'started_at')
    search_fields = ('full_name',)
    readonly_fields = ('id', 'started_at', 'completed_at', 'telegram_sent',
                       'audio_deleted', 'audio_delete_after')
    inlines = [QuestionResponseInline]
    actions = [export_sessions_xlsx]
    date_hierarchy = 'started_at'

    def audio_status(self, obj):
        if obj.audio_deleted:
            return format_html('<span style="color:#888">🗑 Deleted</span>')
        if obj.audio_delete_after:
            return format_html('<span style="color:#f5a623">⏳ Until {}</span>',
                               obj.audio_delete_after.strftime('%m-%d'))
        return format_html('<span style="color:#00d4aa">💾 Saved</span>')
    audio_status.short_description = 'Audio'


@admin.register(SiteAnalytics)
class SiteAnalyticsAdmin(admin.ModelAdmin):
    list_display = ('date', 'total_students', 'new_students', 'total_tests',
                    'total_outsider_tests', 'avg_score')
    readonly_fields = ('date', 'total_students', 'new_students', 'total_tests',
                       'total_outsider_tests', 'avg_score')






# ── Announcements ─────────────────────────────────────────────────────────────

@admin.register(Announcement)
class AnnouncementAdmin(admin.ModelAdmin):
    list_display  = ('emoji', 'title', 'style', 'is_active', 'show_once', 'created_at', 'dismissal_count')
    list_editable = ('is_active',)
    list_filter   = ('style', 'is_active', 'show_once')
    search_fields = ('title', 'body')
    readonly_fields = ('created_at', 'updated_at')
    fieldsets = (
        (None, {'fields': ('emoji', 'title', 'body', 'style')}),
        ('Visibility', {'fields': ('is_active', 'show_once')}),
        ('Timestamps', {'fields': ('created_at', 'updated_at'), 'classes': ('collapse',)}),
    )

    def dismissal_count(self, obj):
        return obj.dismissals.count()
    dismissal_count.short_description = 'Dismissals'
